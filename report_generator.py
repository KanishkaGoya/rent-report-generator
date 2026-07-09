"""
report_generator.py

Core business logic for the Expense Report Generator application.

This module handles:
    - Reading and validating the Vendor Master and expense Report Excel files
    - Cleaning and normalizing data types
    - Grouping and summing duplicate suppliers (allowing positive/negative offset)
    - Merging expense Report data with Vendor Master data
    - Building formatted address strings
    - Generating the final formatted Excel workbook (in-memory, via BytesIO)
    - Generating the Missing_Vendors worksheet for unmatched suppliers

No file is ever written to disk. Everything is handled in-memory so that the
application is safe to run on Streamlit Community Cloud (read-only / ephemeral
filesystem friendly).
"""

from io import BytesIO
from typing import List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# --------------------------------------------------------------------------
# Constants
# --------------------------------------------------------------------------

# Required headers for the Vendor Master upload
VENDOR_MASTER_REQUIRED_COLUMNS = [
    "Vendor",
    "City",
    "Name 1",
    "GST Registration No.",
    "Name 2",
    "Street",
    "Name 3",
    "Name 4",
    "PostalCode",
    "PAN",
]

# Required headers for the Expense Report upload
EXPENSE_REPORT_REQUIRED_COLUMNS = [
    "Supplier",
    "Amount",
]

# Font / style constants (Arial family as specified)
FONT_NAME = "Arial"
COMPANY_HEADING_FONT = Font(name=FONT_NAME, size=14, bold=True)
SUB_HEADING_FONT = Font(name=FONT_NAME, size=13, bold=True)
COLUMN_HEADER_FONT = Font(name=FONT_NAME, size=13, bold=True)
SUPPLIER_NAME_FONT = Font(name=FONT_NAME, size=13, bold=True)
ADDRESS_FONT = Font(name=FONT_NAME, size=13, bold=False)
AMOUNT_FONT = Font(name=FONT_NAME, size=13, bold=False)
VENDOR_CODE_FONT = Font(name=FONT_NAME, size=13, bold=False)
GST_FONT = Font(name=FONT_NAME, size=13, bold=False)

AMOUNT_NUMBER_FORMAT = "#,##0.00"

RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# Fixed column widths used as a fallback since true "AutoFit" is not natively
# supported by openpyxl. These widths are generous enough to accommodate
# typical vendor names, addresses, and amounts without truncation.
COLUMN_WIDTHS = {
    "A": 18,   # Vendor Code
    "B": 55,   # Particulars (Name / Address lines)
    "C": 22,   # GST No
    "D": 20,   # Amount(Rs)
}


# --------------------------------------------------------------------------
# Custom Exceptions
# --------------------------------------------------------------------------

class ExpenseReportError(Exception):
    """Base exception for all user-facing errors raised by this module."""
    pass


class MissingColumnsError(ExpenseReportError):
    """Raised when an uploaded file is missing one or more required columns."""

    def __init__(self, file_label: str, missing_columns: List[str]):
        self.file_label = file_label
        self.missing_columns = missing_columns
        message = (
            f"The '{file_label}' file is missing required column(s): "
            f"{', '.join(missing_columns)}. "
            f"Please check the file and re-upload."
        )
        super().__init__(message)


class EmptyFileError(ExpenseReportError):
    """Raised when an uploaded Excel file contains no data rows."""

    def __init__(self, file_label: str):
        self.file_label = file_label
        message = (
            f"The '{file_label}' file appears to be empty. "
            f"Please upload a file that contains data."
        )
        super().__init__(message)


class InvalidExcelFileError(ExpenseReportError):
    """Raised when an uploaded file cannot be parsed as a valid Excel file."""

    def __init__(self, file_label: str, original_error: str = ""):
        self.file_label = file_label
        message = (
            f"The '{file_label}' file could not be read as a valid Excel "
            f"file. Please make sure it is a valid .xlsx file."
        )
        if original_error:
            message += f" (Details: {original_error})"
        super().__init__(message)


# --------------------------------------------------------------------------
# File Reading & Validation
# --------------------------------------------------------------------------

def read_excel_file(uploaded_file, file_label: str) -> pd.DataFrame:
    """
    Read an uploaded Excel file into a pandas DataFrame.

    For Vendor Master:
        - Reads the first sheet.

    For Expense Report:
        - Searches all sheets and automatically selects the one
          containing the required columns.
    """
    try:
        uploaded_file.seek(0)

        # Expense Report: Search all sheets
        if file_label == "Expense Report":

            excel_file = pd.ExcelFile(uploaded_file, engine="openpyxl")
            dataframe = None

            for sheet_name in excel_file.sheet_names:

                df = pd.read_excel(
                    excel_file,
                    sheet_name=sheet_name,
                    engine="openpyxl",
                )

                df.columns = [
                    str(col).replace("\xa0", " ").strip()
                    for col in df.columns
                ]

                if all(col in df.columns for col in EXPENSE_REPORT_REQUIRED_COLUMNS):
                    dataframe = df
                    break

            if dataframe is None:
                raise MissingColumnsError(
                    file_label,
                    EXPENSE_REPORT_REQUIRED_COLUMNS,
                )

        # Vendor Master: Read first sheet
        else:
            dataframe = pd.read_excel(
                uploaded_file,
                engine="openpyxl",
            )

    except Exception as exc:
        raise InvalidExcelFileError(file_label, str(exc)) from exc

    if dataframe is None or dataframe.empty:
        raise EmptyFileError(file_label)

    dataframe.columns = [
        str(col).replace("\xa0", " ").strip()
        for col in dataframe.columns
    ]

    return dataframe


def validate_required_columns(
    dataframe: pd.DataFrame, required_columns: List[str], file_label: str
) -> None:
    """
    Validate that a DataFrame contains all required columns.

    Args:
        dataframe: The DataFrame to validate.
        required_columns: List of column names that must be present.
        file_label: Human-readable label for the file, used in error
            messages.

    Raises:
        MissingColumnsError: If one or more required columns are missing.
    """
    existing_columns = set(dataframe.columns)
    missing_columns = [
        col for col in required_columns if col not in existing_columns
    ]

    if missing_columns:
        raise MissingColumnsError(file_label, missing_columns)


# --------------------------------------------------------------------------
# Data Cleaning
# --------------------------------------------------------------------------

def _safe_str(value) -> str:
    """
    Safely convert a value to a clean, stripped string.

    Handles NaN/None values by converting them to an empty string, and
    strips any trailing '.0' that pandas sometimes introduces when a
    numeric-looking column (e.g. vendor codes) is read as float.
    """
    if pd.isna(value):
        return ""
    text = str(value).strip()
    # Remove a trailing ".0" artifact caused by float conversion of
    # integer-like vendor/supplier codes (e.g. "100001.0" -> "100001").
    if text.endswith(".0"):
        integer_part = text[:-2]
        if integer_part.replace("-", "", 1).isdigit():
            text = integer_part
    return text


def clean_vendor_master(vendor_master_df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean and normalize the Vendor Master DataFrame.

    - Converts 'Vendor' to string (safely).
    - Strips whitespace from all text fields used downstream.
    - Fills blank/NaN text fields with empty strings.

    Args:
        vendor_master_df: Raw Vendor Master DataFrame.

    Returns:
        A cleaned copy of the DataFrame.
    """
    # Keep only the required columns
    df = vendor_master_df[VENDOR_MASTER_REQUIRED_COLUMNS].copy()

    df["Vendor"] = df["Vendor"].apply(_safe_str)

    text_columns = [
        "City",
        "Name 1",
        "GST Registration No.",
        "PAN",
        "Name 2",
        "Street",
        "Name 3",
        "Name 4",
        "PostalCode",
    ]
    for col in text_columns:
        df[col] = df[col].apply(_safe_str)
    
    # Keep only one record for each Vendor code
    df = df.drop_duplicates(subset=["Vendor"], keep="first")

    return df


def clean_expense_report(expense_report_df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean and normalize the Expense Report DataFrame.

    - Keeps only the 'Supplier' and 'Amount' columns (all other columns are
      ignored per specification).
    - Converts 'Supplier' to string (safely).
    - Converts 'Amount' to numeric, coercing invalid/blank values to 0.

    Args:
        expense_report_df: Raw Expense Report DataFrame.

    Returns:
        A cleaned DataFrame containing only 'Supplier' and 'Amount' columns.
    """
    df = expense_report_df[["Supplier", "Amount"]].copy()

    df["Supplier"] = df["Supplier"].apply(_safe_str)

    # Coerce non-numeric / blank amounts to NaN, then fill with 0 so that
    # blanks are handled safely and do not break the aggregation.
    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce").fillna(0.0)

    # Drop rows where Supplier is completely blank - these cannot be
    # meaningfully merged or reported.
    df = df[df["Supplier"] != ""]

    return df


def aggregate_supplier_amounts(expense_report_df: pd.DataFrame) -> pd.DataFrame:
    """
    Group the Expense Report by Supplier and sum the Amount column.

    Positive and negative amounts for the same supplier naturally offset
    each other during summation. Suppliers whose net amount is zero are
    intentionally NOT removed - they must still appear in the final report.

    Args:
        expense_report_df: Cleaned Expense Report DataFrame with 'Supplier' and
            'Amount' columns.

    Returns:
        A DataFrame with one row per unique Supplier and the summed Amount.
    """
    aggregated_df = (
        expense_report_df.groupby("Supplier", as_index=False)["Amount"]
        .sum()
        .reset_index(drop=True)
    )
    return aggregated_df
# --------------------------------------------------------------------------
# Address Construction
# --------------------------------------------------------------------------

def build_address_lines(row: pd.Series) -> List[str]:
    """
    Build the list of address lines for a given vendor row, in the
    prescribed order, skipping blank values.

    Order:
        1. Name 2
        2. Street
        3. Name 3
        4. Name 4
        5. "City - PostalCode" (or just "City" if PostalCode is blank)

    Args:
        row: A pandas Series representing a single merged vendor/expense row.
            Must contain 'Name 2', 'Street', 'Name 3', 'Name 4', 'City',
            and 'PostalCode'.

    Returns:
        A list of non-empty address line strings, in display order.
    """
    address_lines: List[str] = []

    for field in ["Name 2", "Street", "Name 3", "Name 4"]:
        value = str(row.get(field, "")).strip()
        if value:
            address_lines.append(value)

    city = str(row.get("City", "")).strip()
    postal_code = str(row.get("PostalCode", "")).strip()

    if city and postal_code:
        address_lines.append(f"{city} - {postal_code}")
    elif city:
        address_lines.append(city)
    elif postal_code:
        # Edge case: PostalCode present but City blank. Still show it so no
        # data is silently dropped.
        address_lines.append(postal_code)

    return address_lines


# --------------------------------------------------------------------------
# Excel Report Generation
# --------------------------------------------------------------------------

def _apply_column_widths(worksheet: Worksheet) -> None:
    """Apply fixed, generous column widths to approximate AutoFit behavior."""
    for column_letter, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column_letter].width = width


def _write_report_header(
    worksheet: Worksheet,
    company_name: str,
    assessment_year: str,
    report_title: str,
) -> int:
    """
    Write the three-line report header (company name, assessment year,
    report title), each merged across the four report columns and styled
    per specification.

    Args:
        worksheet: The worksheet to write into.
        company_name: Company name text.
        assessment_year: Assessment year text (e.g. "2025-26").
        report_title: Report title text (e.g. "DETAIL OF EXPENSES EXPENSES").

    Returns:
        The next available (1-indexed) row number after the header block.
    """
    header_lines = [
        company_name.strip().upper(),
        f"ASSESSMENT YEAR {assessment_year.strip()}",
        report_title.strip().upper(),
    ]

    current_row = 1
    for line in header_lines:
        worksheet.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=4
        )
        cell = worksheet.cell(row=current_row, column=1, value=line)
        cell.font = COMPANY_HEADING_FONT
        cell.alignment = CENTER_ALIGN
        current_row += 1

    # Blank row after header block
    current_row += 1

    return current_row


def _write_column_headers(worksheet: Worksheet, start_row: int) -> int:
    """
    Write the column header row (Vendor Code, Particulars, GST No,
    Amount(Rs)).

    Args:
        worksheet: The worksheet to write into.
        start_row: The 1-indexed row number to write the headers on.

    Returns:
        The next available row number after the column header row.
    """
    headers = ["Vendor Code", "Particulars", "GST No", "Amount(Rs)"]
    for col_index, header_text in enumerate(headers, start=1):
        cell = worksheet.cell(row=start_row, column=col_index, value=header_text)
        cell.font = COLUMN_HEADER_FONT
        cell.alignment = (
            RIGHT_ALIGN if header_text == "Amount(Rs)" else LEFT_ALIGN
        )

    return start_row + 1


def _write_supplier_block(
    worksheet: Worksheet, start_row: int, vendor_row: pd.Series
) -> int:
    """
    Write a single supplier's full block into the worksheet: the main data
    row (Vendor Code / Name / GST / Amount) followed by one row per address
    line, then a blank spacer row.

    Args:
        worksheet: The worksheet to write into.
        start_row: The 1-indexed row number to begin writing at.
        vendor_row: A pandas Series with merged vendor + amount data.

    Returns:
        The next available row number after this supplier's block
        (including the trailing blank spacer row).
    """
    vendor_code = str(vendor_row.get("Vendor", "")).strip()
    vendor_name = str(vendor_row.get("Name 1", "")).strip().upper()
    gst_number = str(vendor_row.get("GST Registration No.", "")).strip()
    pan_number = str(vendor_row.get("PAN", "")).strip()

    # Use PAN if GST is blank
    display_tax_id = gst_number if gst_number else pan_number
    amount = vendor_row.get("Amount", 0.0)
    current_row = start_row

    # Main row: Vendor Code | Vendor Name | GST No | Amount
    code_cell = worksheet.cell(row=current_row, column=1, value=vendor_code)
    code_cell.font = VENDOR_CODE_FONT
    code_cell.alignment = LEFT_ALIGN

    name_cell = worksheet.cell(row=current_row, column=2, value=vendor_name)
    name_cell.font = SUPPLIER_NAME_FONT
    name_cell.alignment = LEFT_ALIGN

    gst_cell = worksheet.cell(row=current_row, column=3, value=display_tax_id)
    gst_cell.font = GST_FONT
    gst_cell.alignment = LEFT_ALIGN

    amount_cell = worksheet.cell(row=current_row, column=4, value=float(amount))
    amount_cell.font = AMOUNT_FONT
    amount_cell.alignment = RIGHT_ALIGN
    amount_cell.number_format = AMOUNT_NUMBER_FORMAT

    current_row += 1

    # Address lines - each on its own row, in the "Particulars" column
    address_lines = build_address_lines(vendor_row)
    for address_line in address_lines:
        address_cell = worksheet.cell(row=current_row, column=2, value=address_line)
        address_cell.font = ADDRESS_FONT
        address_cell.alignment = LEFT_ALIGN
        current_row += 1

    # One blank spacer row between suppliers
    current_row += 1

    return current_row


def _build_main_report_sheet(
    workbook: Workbook,
    matched_df: pd.DataFrame,
    company_name: str,
    assessment_year: str,
    report_title: str,
) -> None:
    """
    Build the main "Expense Report" worksheet containing the formatted header,
    column headers, and one block per supplier.

    Args:
        workbook: The openpyxl Workbook to add the sheet to.
        matched_df: DataFrame of matched supplier/vendor rows (with Amount).
        company_name: Company name for the report header.
        assessment_year: Assessment year for the report header.
        report_title: Report title for the report header.
    """
    worksheet = workbook.active
    worksheet.title = "Expense Report"

    current_row = _write_report_header(
        worksheet, company_name, assessment_year, report_title
    )
    current_row = _write_column_headers(worksheet, current_row)

    for _, vendor_row in matched_df.iterrows():
        current_row = _write_supplier_block(worksheet, current_row, vendor_row)

    _apply_column_widths(worksheet)


def _build_missing_vendors_sheet(
    workbook: Workbook, missing_df: pd.DataFrame
) -> None:
    """
    Build the "Missing_Vendors" worksheet listing suppliers present in the
    Expense Report but not found in the Vendor Master.

    Args:
        workbook: The openpyxl Workbook to add the sheet to.
        missing_df: DataFrame with columns ['Supplier', 'Amount'].
    """
    worksheet = workbook.create_sheet(title="Missing_Vendors")

    headers = ["Supplier", "Amount"]
    for col_index, header_text in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=header_text)
        cell.font = COLUMN_HEADER_FONT
        cell.alignment = LEFT_ALIGN if header_text == "Supplier" else RIGHT_ALIGN

    current_row = 2
    for _, row in missing_df.iterrows():
        supplier_cell = worksheet.cell(
            row=current_row, column=1, value=str(row["Supplier"])
        )
        supplier_cell.font = ADDRESS_FONT
        supplier_cell.alignment = LEFT_ALIGN

        amount_cell = worksheet.cell(
            row=current_row, column=2, value=float(row["Amount"])
        )
        amount_cell.font = AMOUNT_FONT
        amount_cell.alignment = RIGHT_ALIGN
        amount_cell.number_format = AMOUNT_NUMBER_FORMAT

        current_row += 1

    worksheet.column_dimensions["A"].width = 25
    worksheet.column_dimensions["B"].width = 20


def generate_expense_report_workbook(
    matched_df: pd.DataFrame,
    missing_df: pd.DataFrame,
    company_name: str,
    assessment_year: str,
    report_title: str,
) -> BytesIO:
    """
    Generate the complete final Excel workbook containing the formatted
    Expense Report sheet and the Missing_Vendors sheet.

    Args:
        matched_df: DataFrame of matched supplier/vendor rows (with Amount).
        missing_df: DataFrame of unmatched suppliers ['Supplier', 'Amount'].
        company_name: Company name for the report header.
        assessment_year: Assessment year for the report header.
        report_title: Report title for the report header.

    Returns:
        A BytesIO buffer containing the generated .xlsx file, positioned at
        the start of the stream and ready to be read/downloaded.
    """
    workbook = Workbook()

    _build_main_report_sheet(
        workbook, matched_df, company_name, assessment_year, report_title
    )
    _build_missing_vendors_sheet(workbook, missing_df)

    output_buffer = BytesIO()
    workbook.save(output_buffer)
    output_buffer.seek(0)

    return output_buffer


# --------------------------------------------------------------------------
# High-Level Orchestration
# --------------------------------------------------------------------------

def process_expense_report(
    vendor_master_file,
    expense_report_file,
    company_name: str,
    assessment_year: str,
    report_title: str,
) -> Tuple[BytesIO, int, int]:
    """
    High-level orchestration function that ties together the full pipeline:
    reading, validating, cleaning, aggregating, merging, and generating the
    final Excel report.

    Args:
        vendor_master_file: Uploaded Vendor Master file (file-like object).
        expense_report_file: Uploaded Expense Report file (file-like object).
        company_name: Company name for the report header.
        assessment_year: Assessment year for the report header.
        report_title: Report title for the report header.

    Returns:
        A tuple of (excel_buffer, matched_count, missing_count):
            excel_buffer: BytesIO buffer containing the final .xlsx file.
            matched_count: Number of suppliers successfully matched.
            missing_count: Number of suppliers not found in Vendor Master.

    Raises:
        ExpenseReportError (or a subclass): For any validation or processing
            failure that should be surfaced to the user as a friendly
            message.
    """
    # Step 1: Read raw files
    vendor_master_raw_df = read_excel_file(vendor_master_file, "Vendor Master")
    expense_report_raw_df = read_excel_file(expense_report_file, "Expense Report")

    # Step 2: Validate required columns
    validate_required_columns(
        vendor_master_raw_df, VENDOR_MASTER_REQUIRED_COLUMNS, "Vendor Master"
    )
    validate_required_columns(
        expense_report_raw_df, EXPENSE_REPORT_REQUIRED_COLUMNS, "Expense Report"
    )

    # Step 3: Clean data
    vendor_master_df = clean_vendor_master(vendor_master_raw_df)
    expense_report_df = clean_expense_report(expense_report_raw_df)

    if expense_report_df.empty:
        raise EmptyFileError("Expense Report")

# Step 4: Aggregate duplicate suppliers
aggregated_expense_df = aggregate_supplier_amounts(expense_report_df)

# Step 5: Remove duplicate Vendor codes from Vendor Master
vendor_lookup = (
    vendor_master_df
    .drop_duplicates(subset=["Vendor"], keep="first")
)

# Lookup Vendor details
matched_df = aggregated_expense_df.merge(
    vendor_lookup,
    left_on="Supplier",
    right_on="Vendor",
    how="left"
)

# Missing suppliers
missing_df = (
    matched_df[matched_df["Vendor"].isna()][["Supplier", "Amount"]]
    .sort_values(by="Amount", ascending=False)
    .reset_index(drop=True)
)

# Matched suppliers
matched_df = (
    matched_df[matched_df["Vendor"].notna()]
    .sort_values(by="Amount", ascending=False)
    .reset_index(drop=True)
)
    # Step 6: Generate the final workbook
    excel_buffer = generate_expense_report_workbook(
    matched_df, missing_df, company_name, assessment_year, report_title
    )

    return excel_buffer, len(matched_df), len(missing_df)
